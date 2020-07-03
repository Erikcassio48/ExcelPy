

import openpyxl, pprint
print('Abrindo a planilha...')
wb = openpyxl.load_workbook('c:\Venv\dados.xlsx')
sheet = wb.get_sheet_by_name('Plan1')
dadosregionais = {} #Criando a chave

row_count = sheet.max_row
column_count = sheet.max_column


print('Lendo as linhas...')
for row in range(2, row_count + 1):
    # Selecionamos as colunas B,C e D e criamos o laço For.
    estados  = sheet['B' + str(row)].value
    produtos = sheet['C' + str(row)].value
    vendas    = sheet['D' + str(row)].value

    # Criamos as chaves para os estados.
    dadosregionais.setdefault(estados, {})
    
   
    dadosregionais[estados].setdefault(produtos, {'Quantidade': 0, 'vendas': 0})

    # Cada linha representa um produto vendido, somamos ao produto
    dadosregionais[estados][produtos]['Quantidade'] += 1

    # Somamos a quantidade de vendas do produto por estado
    dadosregionais[estados][produtos]['vendas'] += int(vendas)

# Abrimos um arquivo vendas.py com o resultado.
print('Escrevendo Resultados')
resultFile = open('vendas.py', 'w')
resultFile.write('allData = ' + pprint.pformat(dadosregionais))
resultFile.close()
print('Done.')
